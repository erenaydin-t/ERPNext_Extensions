# Copyright (c) 2026, Farbod Siyahpoosh and contributors

from __future__ import annotations

import csv
import io
import unittest
from types import SimpleNamespace
from unittest.mock import patch

import frappe
from frappe.utils import cint

from erpnext_extensions.iran_accounting.account_explorer import api, export
from erpnext_extensions.iran_accounting.tests.test_account_explorer_fixtures import (
	build_payload,
	current_fiscal_year,
	disable_export,
	enable_export,
	enable_wave2a_analysis,
	enable_wave2b_voucher,
	enable_wave2c_unified_party,
	require_site,
)


def _fake_spec(view_axis: str, *, dimension_type: str | None = None):
	return SimpleNamespace(
		view_axis=view_axis,
		dimension_scope=SimpleNamespace(dimension_type=dimension_type),
	)


class TestAccountExplorerExportStructure(unittest.TestCase):
	def test_csv_structure(self):
		spec = _fake_spec("account_level")
		columns = export.get_export_columns(spec)
		rows = [
			{
				"display_code": "1000",
				"display_title": "Cash",
				"opening_debit": 10,
				"opening_credit": 0,
				"period_debit": 100,
				"period_credit": 50,
				"debit_balance": 60,
				"credit_balance": 0,
			}
		]
		content = export.build_csv_content(rows, columns)
		reader = csv.reader(io.StringIO(content))
		parsed = list(reader)
		self.assertEqual(len(parsed), 2)
		self.assertIn("Account Code", parsed[0])
		self.assertIn("Closing Credit", parsed[0])
		self.assertEqual(parsed[1][0], "1000")

	def test_xlsx_generation(self):
		spec = _fake_spec("party")
		columns = export.get_export_columns(spec)
		rows = [
			{
				"party_type": "Customer",
				"display_code": "CUST-001",
				"display_title": "Acme",
				"period_debit": 100,
				"period_credit": 0,
				"debit_balance": 100,
				"credit_balance": 0,
			}
		]
		content = export.build_xlsx_content(rows, columns)
		self.assertTrue(content.startswith(b"PK"))

	def test_currency_export_columns_are_native(self):
		columns = export.get_export_columns(_fake_spec("currency"))
		labels = [column["label"] for column in columns]
		self.assertEqual(
			labels,
			[
				"Currency",
				"Debit Amount (Currency)",
				"Debit Amount (Company)",
				"Credit Amount (Currency)",
				"Credit Amount (Company)",
				"Balance (Currency)",
				"Balance (Company)",
			],
		)
		fieldnames = [column["fieldname"] for column in columns]
		self.assertIn("company_period_debit", fieldnames)
		self.assertNotIn("presentation_currency", fieldnames)


class TestAccountExplorerExport(unittest.TestCase):
	def setUp(self):
		self.company = require_site(self)
		enable_wave2a_analysis()
		enable_wave2b_voucher(include_wave2a=False)
		enable_wave2c_unified_party(include_wave2b=True)
		enable_export()
		frappe.set_user("Administrator")
		fy = current_fiscal_year(self.company)
		if not fy:
			self.skipTest("No fiscal year")
		self.fiscal_year, self.from_date, self.to_date = fy

	def tearDown(self):
		frappe.set_user("Administrator")
		enable_export(threshold=5000)

	def _reset_response(self):
		frappe.local.response = frappe._dict()

	def _export_sync(self, payload, file_format="csv"):
		self._reset_response()
		return export.export_account_explorer(payload, file_format, force_sync=True)

	def test_export_disabled_permission(self):
		disable_export()
		payload = build_payload(self.company, self.fiscal_year, self.from_date, self.to_date)
		with self.assertRaises(frappe.ValidationError):
			self._export_sync(payload)
		enable_export()

	def test_metadata_includes_export_settings(self):
		meta = api.get_metadata()
		self.assertIn("export_enabled", meta)
		self.assertIn("export_background_threshold", meta)
		self.assertEqual(meta.get("export_enabled"), 1)

	def test_account_axis_export(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		result = self._export_sync(payload, "csv")
		self.assertEqual(result.get("queued"), 0)
		self.assertEqual(frappe.local.response.type, "download")
		reader = csv.reader(io.StringIO(frappe.local.response.filecontent))
		headers = next(reader)
		self.assertIn("Account Code", headers)
		self.assertIn("Opening Debit", headers)
		self.assertIn("Closing Credit", headers)

	def test_party_axis_export(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			analysis={"view_axis": "party"},
			document={"hide_zero_rows": 0},
		)
		self._export_sync(payload, "csv")
		reader = csv.reader(io.StringIO(frappe.local.response.filecontent))
		headers = next(reader)
		self.assertIn("Party Type", headers)
		self.assertIn("Party Name", headers)

	def test_dimension_axis_export(self):
		if not frappe.get_meta("GL Entry").has_field("cost_center"):
			self.skipTest("No cost_center on GL Entry")
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			analysis={
				"view_axis": "dimension",
				"dimension_scope": {"dimension_type": "cost_center"},
			},
			document={"hide_zero_rows": 0},
		)
		self._export_sync(payload, "csv")
		reader = csv.reader(io.StringIO(frappe.local.response.filecontent))
		headers = next(reader)
		self.assertIn("Dimension Type", headers)
		self.assertIn("Dimension Value", headers)

	def test_currency_export_without_conversion(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			analysis={"view_axis": "currency"},
			document={"hide_zero_rows": 0},
		)
		self._export_sync(payload, "csv")
		reader = csv.reader(io.StringIO(frappe.local.response.filecontent))
		headers = next(reader)
		company_currency = frappe.get_cached_value("Company", self.company, "default_currency") or "Company"
		self.assertEqual(
			headers,
			[
				"Currency",
				"Debit Amount (Currency)",
				f"Debit Amount ({company_currency})",
				"Credit Amount (Currency)",
				f"Credit Amount ({company_currency})",
				"Balance (Currency)",
				f"Balance ({company_currency})",
			],
		)

	def test_voucher_axis_export(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			analysis={"view_axis": "voucher"},
			document={"hide_zero_rows": 0},
		)
		self._export_sync(payload, "xlsx")
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filecontent.startswith(b"PK"))
		self.assertTrue(frappe.local.response.filename.endswith(".xlsx"))

	def test_company_permission(self):
		payload = build_payload(self.company, self.fiscal_year, self.from_date, self.to_date)
		with patch(
			"erpnext_extensions.iran_accounting.account_explorer.query_spec.assert_company_allowed",
			side_effect=frappe.PermissionError("Not permitted for company"),
		):
			with self.assertRaises(frappe.PermissionError):
				self._export_sync(payload)

	def test_large_export_threshold(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		# Explicit positive threshold; values < 1 fall back to DEFAULT and must not queue.
		enable_export(threshold=1)
		with patch("frappe.enqueue") as enqueue_mock:
			with patch.object(export, "_probe_export_size", return_value=2):
				result = export.export_account_explorer(payload, "csv", force_sync=False)
		self.assertEqual(result.get("queued"), 1)
		enqueue_mock.assert_called_once()
		self.assertIn("background", (result.get("message") or "").lower())

	def test_e01_threshold_zero_does_not_queue_small_export(self):
		"""E01 / E09: stored threshold 0 must normalize to 5000 — small sets stay sync."""
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		enable_export(threshold=0)
		self.assertEqual(export.normalize_export_background_threshold(0), 5000)
		self.assertEqual(export.normalize_export_background_threshold(None), 5000)
		self.assertEqual(export._export_settings()["export_background_threshold"], 5000)
		self._reset_response()
		with patch("frappe.enqueue") as enqueue_mock:
			with patch.object(export, "_probe_export_size", return_value=10):
				with patch.object(export, "collect_export_rows", return_value=([], {}, 10)):
					result = export.export_account_explorer(payload, "xlsx", force_sync=False)
		self.assertNotEqual(result.get("queued"), 1)
		enqueue_mock.assert_not_called()
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filecontent.startswith(b"PK"))

	def test_e09_metadata_threshold_never_zero(self):
		enable_export(threshold=0)
		meta = api.get_metadata()
		self.assertGreaterEqual(cint(meta.get("export_background_threshold")), 1)
		self.assertEqual(cint(meta.get("export_background_threshold")), 5000)

	def test_e10_explicit_configured_threshold_respected(self):
		enable_export(threshold=7)
		self.assertEqual(export.normalize_export_background_threshold(7), 7)
		meta = api.get_metadata()
		self.assertEqual(cint(meta.get("export_background_threshold")), 7)

	def test_e11_e12_threshold_boundary(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		enable_export(threshold=10)
		self._reset_response()
		with patch("frappe.enqueue") as enqueue_mock:
			with patch.object(export, "_probe_export_size", return_value=10):
				with patch.object(export, "collect_export_rows", return_value=([], {}, 10)):
					below_or_equal = export.export_account_explorer(payload, "csv", force_sync=False)
			with patch.object(export, "_probe_export_size", return_value=11):
				above = export.export_account_explorer(payload, "csv", force_sync=False)
		self.assertNotEqual(below_or_equal.get("queued"), 1)
		self.assertEqual(above.get("queued"), 1)
		self.assertEqual(enqueue_mock.call_count, 1)

	def test_e13_queued_response_is_structured(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		enable_export(threshold=1)
		self._reset_response()
		with patch("frappe.enqueue"):
			with patch.object(export, "_probe_export_size", return_value=5):
				result = export.export_account_explorer(payload, "xlsx", force_sync=False)
		self.assertEqual(result.get("queued"), 1)
		self.assertIn("total_rows", result)
		self.assertIn("message", result)
		self.assertNotEqual(frappe.local.response.get("type"), "download")

	def test_e02_e05_e06_xlsx_unicode_and_numeric(self):
		"""E02–E06: XLSX is valid, Unicode survives, numerics stay numeric."""
		spec = _fake_spec("account_level")
		columns = export.get_export_columns(spec)
		rows = [
			{
				"display_code": "1101",
				"display_title": "حساب نقد و بانک",
				"opening_debit": 10.5,
				"opening_credit": 0,
				"period_debit": 100,
				"period_credit": 50,
				"debit_balance": 60.5,
				"credit_balance": 0,
			}
		]
		content = export.build_xlsx_content(rows, columns)
		self.assertTrue(content.startswith(b"PK"))

		from openpyxl import load_workbook

		wb = load_workbook(io.BytesIO(content))
		ws = wb.active
		matrix = [[cell.value for cell in row] for row in ws.iter_rows()]
		self.assertGreaterEqual(len(matrix), 2)
		self.assertIn("Account Code", matrix[0])
		self.assertEqual(matrix[1][0], "1101")
		self.assertEqual(matrix[1][1], "حساب نقد و بانک")
		debit_idx = matrix[0].index("Period Debit")
		self.assertEqual(matrix[1][debit_idx], 100)
		self.assertIsInstance(matrix[1][debit_idx], (int, float))

	def test_e07_csv_still_works(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		self._export_sync(payload, "csv")
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filename.endswith(".csv"))
		text = frappe.local.response.filecontent
		if isinstance(text, bytes):
			text = text.decode("utf-8")
		self.assertIn(",", text.splitlines()[0])

	def test_e08_empty_result_set_safe(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		with patch.object(export, "collect_export_rows", return_value=([], {}, 0)):
			with patch.object(export, "_probe_export_size", return_value=0):
				self._reset_response()
				export.export_account_explorer(payload, "xlsx", force_sync=True)
		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filecontent.startswith(b"PK"))

	def test_unified_party_axis_export(self):
		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			analysis={"view_axis": "unified_party"},
			document={"hide_zero_rows": 0},
		)
		self._export_sync(payload, "csv")
		reader = csv.reader(io.StringIO(frappe.local.response.filecontent))
		headers = next(reader)
		self.assertIn("Unified Party", headers)
		self.assertIn("Member Count", headers)

	def test_whitelisted_export_entry_point(self):
		from erpnext_extensions.iran_accounting.account_explorer import export_account_explorer

		payload = build_payload(
			self.company,
			self.fiscal_year,
			self.from_date,
			self.to_date,
			document={"hide_zero_rows": 0},
		)
		self._reset_response()
		export_account_explorer(payload, "csv")
		self.assertEqual(frappe.local.response.type, "download")
