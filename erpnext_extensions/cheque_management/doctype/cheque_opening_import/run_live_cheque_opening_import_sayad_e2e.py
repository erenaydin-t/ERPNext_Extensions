"""Live E2E: Cheque Opening Import Sayad Code + Sayad Registered (real COI flow).

bench --site development.localhost execute \\
  erpnext_extensions.cheque_management.doctype.cheque_opening_import.run_live_cheque_opening_import_sayad_e2e.run
"""

from __future__ import annotations

import io
import json
import time
from contextlib import contextmanager
from datetime import timedelta
from typing import Any

import frappe
from frappe.utils import cint, getdate, today

from erpnext_extensions.cheque_management.doctype.cheque_opening_import.cheque_opening_import import (
	TEMPLATE_HEADERS,
)
from erpnext_extensions.cheque_management.doctype.post_dated_cheque.post_dated_cheque import (
	_get_pdc_settings_for_company,
)
from erpnext_extensions.cheque_management.run_live_party_orchestration_e2e import _site_context


def _counts_for_pdc(pdc_name: str | None) -> dict:
	if not pdc_name:
		return {"journal_reference_count": 0, "journal_entry_count": 0}
	n_ref = frappe.db.sql(
		"SELECT COUNT(*) FROM `tabPDC Journal Reference` WHERE parent = %s",
		(pdc_name,),
	)[0][0]
	return {"journal_reference_count": n_ref, "journal_entry_count": 0 if n_ref == 0 else n_ref}


def _ensure_pdc_settings_for_company(company: str) -> None:
	"""Link orphaned PDC Settings to the company used by opening-import E2E (site hygiene)."""
	if _get_pdc_settings_for_company(company):
		return
	orphan = frappe.db.get_value(
		"PDC Settings",
		{"default_cheques_in_hand_account": ("is", "set")},
		"name",
		order_by="modified desc",
	)
	if not orphan:
		return
	if not frappe.db.exists("Company", company):
		return
	frappe.db.set_value("PDC Settings", orphan, "company", company)
	frappe.db.commit()


def _site_ctx() -> dict:
	snap = frappe.db.sql(
		"""
		SELECT company, party, drawer_bank_name
		FROM `tabPost Dated Cheque`
		WHERE is_opening_import = 1
		  AND cheque_direction = 'Receivable'
		  AND IFNULL(account_paid_to, '') != ''
		ORDER BY creation DESC
		LIMIT 1
		""",
		as_dict=True,
	)
	if snap:
		row = snap[0]
		drawer = (row.get("drawer_bank_name") or "").strip() or frappe.db.get_value(
			"Bank", {}, "name", order_by="modified desc"
		)
		company = row["company"]
		_ensure_pdc_settings_for_company(company)
		return {"company": company, "customer": row["party"], "drawer_bank": drawer}

	base = _site_context()
	drawer_bank = frappe.db.get_value("Bank", {}, "name", order_by="modified desc")
	if not drawer_bank:
		frappe.throw("No Bank master for drawer_bank_name")
	return {
		"company": base["company"],
		"customer": base["customer"],
		"drawer_bank": drawer_bank,
	}


@contextmanager
def _sayad_not_required_for_import(company: str):
	"""Test C: empty Sayad is valid only when PDC Settings do not require Sayad."""
	settings = _get_pdc_settings_for_company(company)
	if not settings:
		yield {"temporarily_relaxed": False}
		return
	settings_name = settings.name
	orig = cint(settings.require_sayad_registration)
	if not orig:
		yield {"temporarily_relaxed": False}
		return
	frappe.db.set_value("PDC Settings", settings_name, "require_sayad_registration", 0)
	frappe.db.commit()
	try:
		yield {"temporarily_relaxed": True, "pdc_settings": settings_name}
	finally:
		frappe.db.set_value("PDC Settings", settings_name, "require_sayad_registration", 1)
		frappe.db.commit()


def _base_receivable_row(
	ctx: dict,
	cheque_no: str,
	workflow_state: str,
	sayad_code: str | None,
	sayad_registered: Any,
) -> list[Any]:
	t0 = getdate(today())
	due = t0 + timedelta(days=30)
	return [
		"Receivable",
		ctx["company"],
		"",
		"",
		cheque_no,
		due,
		"100",
		"Customer",
		ctx["customer"],
		workflow_state,
		ctx["drawer_bank"],
		t0 if workflow_state != "Draft" else None,
		None,
		None,
		None,
		None,
		None,
		sayad_code if sayad_code is not None else "",
		sayad_registered,
	]


def _run_coi_import(
	*,
	test_id: str,
	headers: list[str],
	row: list[Any],
) -> dict:
	import openpyxl

	suffix = f"{test_id}-{int(time.time())}"
	fname = f"coi_sayad_e2e_{suffix}.xlsx"
	site_path = frappe.get_site_path("private", "files", fname)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Data"
	ws.append(headers)
	ws.append(row)
	wb.save(site_path)

	file_url = f"/private/files/{fname}"
	frappe.get_doc(
		{
			"doctype": "File",
			"file_name": fname,
			"file_url": file_url,
			"is_private": 1,
			"folder": "Home",
		}
	).insert(ignore_permissions=True)
	frappe.db.commit()

	coi = frappe.get_doc({"doctype": "Cheque Opening Import", "import_file": file_url})
	coi.insert(ignore_permissions=True)
	frappe.db.commit()

	preview = coi.preview_file()
	frappe.db.commit()
	coi.reload()

	if (coi.import_status or "").strip() != "Previewed":
		return {
			"test_id": test_id,
			"passed": False,
			"phase": "preview",
			"coi_name": coi.name,
			"import_status": coi.import_status,
			"preview": preview,
		}

	execute = coi.execute_import()
	frappe.db.commit()
	coi.reload()

	pdc_name = None
	row_status = None
	for item in coi.items or []:
		if (item.row_status or "") == "Imported":
			pdc_name = item.imported_pdc or item.post_dated_cheque
			row_status = item.row_status
			break
		if not row_status:
			row_status = item.row_status

	sayad = None
	if pdc_name:
		sayad = frappe.db.get_value(
			"Post Dated Cheque",
			pdc_name,
			["sayad_code", "sayad_registered", "workflow_state"],
			as_dict=True,
		)

	counts = _counts_for_pdc(pdc_name)

	return {
		"test_id": test_id,
		"coi_name": coi.name,
		"pdc_name": pdc_name,
		"row_status": row_status,
		"import_status": coi.import_status,
		"preview": preview,
		"execute": execute,
		"sayad_code": (sayad or {}).get("sayad_code"),
		"sayad_registered": cint((sayad or {}).get("sayad_registered")),
		"workflow_state": (sayad or {}).get("workflow_state"),
		"je_count": counts["journal_entry_count"],
		"pdc_journal_reference_count": counts["journal_reference_count"],
	}


def _assert_case(case: dict, *, expect_code: str | None, expect_reg: int) -> list[str]:
	errors = []
	if (case.get("import_status") or "") not in ("Completed",):
		errors.append(f"{case['test_id']}: import_status={case.get('import_status')!r}")
	if case.get("row_status") != "Imported":
		errors.append(f"{case['test_id']}: row_status={case.get('row_status')!r}")
	if not case.get("pdc_name"):
		errors.append(f"{case['test_id']}: missing PDC")
	code = (case.get("sayad_code") or "").strip()
	if expect_code is None:
		if code:
			errors.append(f"{case['test_id']}: expected empty sayad_code, got {code!r}")
	else:
		if code != expect_code:
			errors.append(f"{case['test_id']}: sayad_code expected {expect_code!r}, got {code!r}")
	if cint(case.get("sayad_registered")) != expect_reg:
		errors.append(
			f"{case['test_id']}: sayad_registered expected {expect_reg}, got {case.get('sayad_registered')}"
		)
	if case.get("je_count") != 0 or case.get("pdc_journal_reference_count") != 0:
		errors.append(f"{case['test_id']}: expected 0 JE/refs, got {case}")
	return errors


def _template_headers_test() -> dict:
	import openpyxl

	from frappe.utils import cstr

	from erpnext_extensions.cheque_management.doctype.cheque_opening_import import (
		cheque_opening_import as coi_mod,
	)

	orig_response = frappe.local.response
	frappe.local.response = frappe._dict()
	try:
		coi_mod.download_import_template()
		content = frappe.local.response.filecontent
	finally:
		frappe.local.response = orig_response

	wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
	ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
	headers = [cstr(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
	has_code = "sayad_code" in headers
	has_reg = "sayad_registered" in headers
	return {
		"test_id": "E",
		"passed": has_code and has_reg,
		"headers": headers,
		"has_sayad_code": has_code,
		"has_sayad_registered": has_reg,
	}


def run():
	frappe.set_user("Administrator")
	ctx = _site_ctx()
	base_ts = int(time.time())
	errors: list[str] = []
	cases: list[dict] = []

	# Test A — English headers, Registered, explicit code + registered=1
	headers_a = list(TEMPLATE_HEADERS)
	headers_a[headers_a.index("sayad_code")] = "Sayad Code"
	headers_a[headers_a.index("sayad_registered")] = "Sayad Registered"
	case_a = _run_coi_import(
		test_id="A",
		headers=headers_a,
		row=_base_receivable_row(ctx, f"OB-SAYAD-A-{base_ts}", "Registered", "OB-SAYAD-TEST-001", 1),
	)
	cases.append(case_a)
	errors.extend(_assert_case(case_a, expect_code="OB-SAYAD-TEST-001", expect_reg=1))

	# Test B — code filled, registered column empty → default 1 (Draft avoids extra validation)
	case_b = _run_coi_import(
		test_id="B",
		headers=list(TEMPLATE_HEADERS),
		row=_base_receivable_row(ctx, f"OB-SAYAD-B-{base_ts}", "Draft", "OB-SAYAD-TEST-002", ""),
	)
	cases.append(case_b)
	errors.extend(_assert_case(case_b, expect_code="OB-SAYAD-TEST-002", expect_reg=1))

	# Test C — both empty → registered 0 (requires Sayad optional per PDC Settings on insert)
	with _sayad_not_required_for_import(ctx["company"]) as relax:
		case_c = _run_coi_import(
			test_id="C",
			headers=list(TEMPLATE_HEADERS),
			row=_base_receivable_row(ctx, f"OB-SAYAD-C-{base_ts}", "Draft", "", ""),
		)
		case_c["sayad_policy_note"] = (
			"PDC Settings require_sayad_registration temporarily disabled for this case only"
			if relax.get("temporarily_relaxed")
			else "PDC Settings already allow optional Sayad"
		)
	cases.append(case_c)
	errors.extend(_assert_case(case_c, expect_code=None, expect_reg=0))

	# Test D — code filled, explicit no (Persian خیر) → registered 0
	case_d = _run_coi_import(
		test_id="D",
		headers=list(TEMPLATE_HEADERS),
		row=_base_receivable_row(ctx, f"OB-SAYAD-D-{base_ts}", "Draft", "OB-SAYAD-TEST-D", "خیر"),
	)
	cases.append(case_d)
	errors.extend(_assert_case(case_d, expect_code="OB-SAYAD-TEST-D", expect_reg=0))

	case_e = _template_headers_test()
	cases.append(case_e)
	if not case_e.get("passed"):
		errors.append("E: generated template missing sayad_code or sayad_registered column")

	pdc_names = [c.get("pdc_name") for c in cases if c.get("pdc_name")]
	sql_evidence = {}
	if pdc_names:
		sql_evidence["pdcs"] = frappe.db.sql(
			"""
			SELECT name, sayad_code, sayad_registered, workflow_state, opening_import
			FROM `tabPost Dated Cheque`
			WHERE name IN ({})
			""".format(", ".join(["%s"] * len(pdc_names))),
			tuple(pdc_names),
			as_dict=True,
		)
		sql_evidence["journal_refs"] = frappe.db.sql(
			"""
			SELECT parent, COUNT(*) AS cnt
			FROM `tabPDC Journal Reference`
			WHERE parent IN ({})
			GROUP BY parent
			""".format(", ".join(["%s"] * len(pdc_names))),
			tuple(pdc_names),
			as_dict=True,
		)

	out = {
		"passed": not errors,
		"errors": errors,
		"site": frappe.local.site,
		"cases": cases,
		"sql_evidence": sql_evidence,
	}
	print(json.dumps(out, indent=2, default=str))
	if errors:
		raise frappe.ValidationError("Sayad COI E2E failed:\n" + "\n".join(errors))
	return out
