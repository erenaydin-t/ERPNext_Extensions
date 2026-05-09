# Copyright (c) 2026, Farbod Siyahpoosh and contributors
# License: MIT. See license.txt

"""Opening balance import for **Post Dated Cheque** (Payable + Receivable).

Uses normal PDC insert/save/submit and workflow transitions — no bypass of validation,
reservation, or leaf Used hooks. Each data row runs in a DB savepoint so failures
do not disturb other rows.
"""

from __future__ import annotations

import io
import json
import re
from collections import deque
from datetime import date
from typing import Any

import frappe
from frappe.database.database import savepoint
from frappe.model.document import Document
from frappe.utils import cstr, flt, getdate

from erpnext_extensions.cheque_management.pdc_workflow_state_machine import (
	ALL_WORKFLOW_STATES,
	CHEQUE_DIRECTION_PAYABLE,
	CHEQUE_DIRECTION_RECEIVABLE,
	PAYABLE_WORKFLOW_TRANSITIONS,
	RECEIVABLE_WORKFLOW_TRANSITIONS,
	normalize_workflow_state_value,
)

REQUIRED_HEADERS = frozenset(
	{
		"cheque_direction",
		"company",
		"bank_account",
		"cheque_number",
		"cheque_due_date",
		"cheque_amount",
		"party_type",
		"party",
		"workflow_state",
	}
)

HEADER_ALIASES = {
	"direction": "cheque_direction",
	"cheque_no": "cheque_number",
	"amount": "cheque_amount",
	"due_date": "cheque_due_date",
	"workflow": "workflow_state",
	"status": "workflow_state",
	"book": "cheque_book",
	"drawer_bank": "drawer_bank_name",
	"sayad": "sayad_code",
}


def _norm_header(h: str | None) -> str:
	if h is None:
		return ""
	s = cstr(h).strip().lower()
	s = re.sub(r"[\s\-]+", "_", s)
	return HEADER_ALIASES.get(s, s)


def _parse_date(val: Any):
	if val is None or val == "":
		return None
	if hasattr(val, "date"):
		return val.date()
	return getdate(val)


def _rows_from_sheet(rows: list[list[Any]]) -> list[tuple[int, dict[str, Any]]]:
	if not rows:
		return []
	header = [_norm_header(cstr(c) if c is not None else "") for c in rows[0]]
	missing = [h for h in REQUIRED_HEADERS if h not in header]
	if missing:
		frappe.throw(
			frappe._("Missing required column(s): {0}").format(", ".join(sorted(missing))),
			title=frappe._("Cheque Opening Import"),
		)
	out = []
	for i, line in enumerate(rows[1:], start=2):
		if not line or all(c is None or cstr(c).strip() == "" for c in line):
			continue
		row = {}
		for hi, key in enumerate(header):
			if not key:
				continue
			val = line[hi] if hi < len(line) else None
			if val is not None and not isinstance(val, (str, int, float)):
				val = cstr(val)
			row[key] = val
		out.append((i, row))
	return out


def _get_sheet_from_file_url(file_url: str) -> list[list[Any]]:
	if not file_url:
		frappe.throw(frappe._("Attach a CSV or Excel file first."), title=frappe._("Cheque Opening Import"))

	from frappe.utils.csvutils import read_csv_content
	from frappe.utils.xlsxutils import read_xls_file_from_attached_file

	_file = frappe.get_doc("File", {"file_url": file_url})
	path = _file.get_full_path() or ""
	lower = path.lower()

	logger = frappe.logger("cheque_opening_import")
	logger.info("Preview parse start: file_url=%s path=%s", file_url, path)

	if lower.endswith(".csv"):
		content = _file.get_content()
		logger.info("Preview parse: detected_type=csv")
		return read_csv_content(content)

	if lower.endswith(".xlsx"):
		logger.info("Preview parse: detected_type=xlsx")
		import openpyxl

		wb = openpyxl.load_workbook(filename=path, data_only=True)
		ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
		if ws.title != "Data":
			logger.info("Preview parse: sheet_selected=%s (fallback)", ws.title)
		else:
			logger.info("Preview parse: sheet_selected=Data")

		rows: list[list[Any]] = []
		for row in ws.iter_rows():
			rows.append([cell.value for cell in row])
		logger.info("Preview parse: parsed_rows=%s", len(rows))
		return rows

	if lower.endswith(".xls"):
		logger.info("Preview parse: detected_type=xls")
		content = _file.get_content()
		rows = read_xls_file_from_attached_file(content)
		logger.info("Preview parse: parsed_rows=%s", len(rows))
		return rows

	frappe.throw(
		frappe._("Unsupported file type. Use .csv or .xlsx."),
		title=frappe._("Cheque Opening Import"),
	)


def find_workflow_path(cheque_direction: str, target_raw: str) -> list[str]:
	"""Shortest path from **Draft** to ``target`` on the direction-specific graph."""
	target = normalize_workflow_state_value(target_raw)
	if target not in ALL_WORKFLOW_STATES:
		frappe.throw(
			frappe._("Invalid workflow_state: {0}").format(target_raw),
			title=frappe._("Cheque Opening Import"),
		)

	graph = (
		RECEIVABLE_WORKFLOW_TRANSITIONS
		if (cheque_direction or "").strip() == CHEQUE_DIRECTION_RECEIVABLE
		else PAYABLE_WORKFLOW_TRANSITIONS
	)
	start = normalize_workflow_state_value(None)  # Draft
	if target == start:
		return [start]

	q: deque[str] = deque([start])
	prev: dict[str, str | None] = {start: None}
	while q:
		u = q.popleft()
		for v in graph.get(u, frozenset()):
			if v not in prev:
				prev[v] = u
				if v == target:
					path: list[str] = []
					cur: str | None = v
					while cur is not None:
						path.append(cur)
						cur = prev[cur]
					path.reverse()
					return path
				q.append(v)

	frappe.throw(
		frappe._("No allowed workflow path from Draft to {0} for {1}.").format(target, cheque_direction),
		title=frappe._("Cheque Opening Import"),
	)


def _resolve_cheque_leaf(row: dict[str, Any]) -> str | None:
	if (row.get("cheque_direction") or "").strip() != CHEQUE_DIRECTION_PAYABLE:
		return None
	book = cstr(row.get("cheque_book") or "").strip()
	if not book:
		return None
	chq = cstr(row.get("cheque_number") or "").strip()
	company = cstr(row.get("company") or "").strip()
	bank = cstr(row.get("bank_account") or "").strip()
	if not chq or not company or not bank:
		return None
	name = frappe.db.get_value(
		"Cheque Leaf",
		{
			"cheque_book": book,
			"cheque_number": chq,
			"company": company,
			"bank_account": bank,
		},
		"name",
	)
	return name or None


def _validate_row_preview(row: dict[str, Any], row_no: int) -> str | None:
	"""Return error message or None if OK."""
	d = (row.get("cheque_direction") or "").strip()
	if d not in (CHEQUE_DIRECTION_PAYABLE, CHEQUE_DIRECTION_RECEIVABLE):
		return f"Row {row_no}: invalid cheque_direction {d!r}"

	for fn in ("company", "party_type", "party", "workflow_state"):
		if not cstr(row.get(fn) or "").strip():
			return f"Row {row_no}: {fn} is required"

	if not cstr(row.get("cheque_number") or "").strip():
		return f"Row {row_no}: cheque_number is required"

	try:
		_parse_date(row.get("cheque_due_date"))
	except Exception:
		return f"Row {row_no}: invalid cheque_due_date"

	try:
		flt(row.get("cheque_amount"))
	except Exception:
		return f"Row {row_no}: invalid cheque_amount"

	if d == CHEQUE_DIRECTION_RECEIVABLE and not cstr(row.get("drawer_bank_name") or "").strip():
		return f"Row {row_no}: drawer_bank_name is required for Receivable"

	ws = cstr(row.get("workflow_state") or "").strip()
	try:
		find_workflow_path(d, ws)
	except Exception as e:
		return f"Row {row_no}: workflow — {e!s}"

	target = normalize_workflow_state_value(ws)
	bank = cstr(row.get("bank_account") or "").strip()

	# Payable: always requires bank_account.
	if d == CHEQUE_DIRECTION_PAYABLE and not bank:
		return f"Row {row_no}: bank_account is required"

	# Receivable: bank_account is required only when "Sent to Bank or later".
	# Keep this aligned with PDC runtime behavior.
	receivable_needs_bank_states_raw = {
		"Sent to Bank",
		"In Clearing",
		"Cleared",
		"Bounced",
		"Returned",
		"Cancelled",
		"Replaced",
		"Under Legal Action",
	}
	receivable_needs_bank = {
		normalize_workflow_state_value(s)
		for s in receivable_needs_bank_states_raw
		if normalize_workflow_state_value(s) in ALL_WORKFLOW_STATES
	}
	if d == CHEQUE_DIRECTION_RECEIVABLE and target in receivable_needs_bank and not bank:
		return f"Row {row_no}: bank_account required for this workflow state"

	leaf = _resolve_cheque_leaf(row)
	if d == CHEQUE_DIRECTION_PAYABLE and cstr(row.get("cheque_book") or "").strip() and not leaf:
		return f"Row {row_no}: Cheque Leaf not found for book + cheque_number + company + bank_account"

	return None


def _prepare_state_dates(pdc: Document, to_state: str, row: dict[str, Any]) -> None:
	"""Set dates/fields typically required when *entering* ``to_state`` (best-effort from row)."""
	ts = normalize_workflow_state_value(to_state)
	rd = _parse_date(row.get("received_date"))
	if rd and ts in (
		normalize_workflow_state_value("Registered"),
		normalize_workflow_state_value("Issued"),
		normalize_workflow_state_value("Sent to Bank"),
		normalize_workflow_state_value("Cleared"),
	):
		if not getattr(pdc, "received_date", None):
			pdc.received_date = rd

	if ts == normalize_workflow_state_value("Issued") and pdc.cheque_direction == CHEQUE_DIRECTION_PAYABLE:
		hd = _parse_date(row.get("handover_date"))
		if hd:
			pdc.handover_date = hd
		elif not getattr(pdc, "handover_date", None):
			pdc.handover_date = getattr(pdc, "cheque_due_date", None)

	if ts == normalize_workflow_state_value("Sent to Bank"):
		sd = _parse_date(row.get("sent_to_bank_date"))
		if sd:
			pdc.sent_to_bank_date = sd

	if ts == normalize_workflow_state_value("Cleared"):
		cd = _parse_date(row.get("cleared_date"))
		if cd:
			pdc.cleared_date = cd

	if ts == normalize_workflow_state_value("Bounced"):
		bd = _parse_date(row.get("bounced_date"))
		if bd:
			pdc.bounced_date = bd

	if ts == normalize_workflow_state_value("Returned"):
		rdt = _parse_date(row.get("returned_date"))
		if rdt:
			pdc.returned_date = rdt


def import_row(row_no: int, row: dict[str, Any]) -> str:
	"""Create PDC, walk workflow, optionally submit. Caller wraps in savepoint."""
	direction = (row.get("cheque_direction") or "").strip()
	target = normalize_workflow_state_value(row.get("workflow_state"))
	path = find_workflow_path(direction, target)

	leaf = _resolve_cheque_leaf(row)
	chq_num = cstr(row.get("cheque_number") or "").strip()
	company = cstr(row.get("company") or "").strip()
	party = cstr(row.get("party") or "").strip()

	# Duplicate guard: prevent re-import of the same opening cheque.
	existing = frappe.db.get_value(
		"Post Dated Cheque",
		{
			"cheque_direction": direction,
			"cheque_no": chq_num,
			"company": company,
			"party": party,
			"is_opening_import": 1,
		},
		"name",
	)
	if existing:
		frappe.throw(frappe._("Opening cheque already imported: {0}").format(existing))

	pdc = frappe.new_doc("Post Dated Cheque")
	pdc.cheque_direction = direction
	pdc.company = company
	pdc.bank_account = cstr(row.get("bank_account") or "").strip()
	pdc.cheque_due_date = _parse_date(row.get("cheque_due_date"))
	pdc.cheque_amount = flt(row.get("cheque_amount"))
	pdc.party_type = cstr(row.get("party_type") or "").strip()
	pdc.party = party
	pdc.workflow_state = normalize_workflow_state_value(None)
	pdc.allocation_mode = "direct_settlement"
	pdc.cheque_no = chq_num

	# Opening import tracking (traceability + re-import protection).
	pdc.is_opening_import = 1
	pdc.opening_import = frappe.flags.get("cheque_opening_import_name")
	pdc.opening_import_row = row_no

	if direction == CHEQUE_DIRECTION_RECEIVABLE:
		pdc.drawer_bank_name = cstr(row.get("drawer_bank_name") or "").strip()

	if leaf:
		pdc.cheque_leaf = leaf
		leaf_no = frappe.db.get_value("Cheque Leaf", leaf, "cheque_number")
		if leaf_no:
			pdc.cheque_no = leaf_no

	sc = cstr(row.get("sayad_code") or "").strip()
	if sc:
		pdc.sayad_code = sc

	rd0 = _parse_date(row.get("received_date"))
	if rd0:
		pdc.received_date = rd0
	elif len(path) > 1:
		pdc.received_date = pdc.cheque_due_date

	pdc.insert()

	for i in range(1, len(path)):
		to_st = path[i]
		_prepare_state_dates(pdc, to_st, row)
		pdc.workflow_state = to_st
		pdc.save()

	final = normalize_workflow_state_value(pdc.workflow_state)
	if final != normalize_workflow_state_value("Draft"):
		pdc.submit()

	return pdc.name


class ChequeOpeningImport(Document):
	@frappe.whitelist()
	def preview(self):
		"""Alias for client-side `frm.call("preview")`."""
		return self.preview_file()

	@frappe.whitelist()
	def preview_file(self):
		self.check_permission("write")
		if not self.name:
			frappe.throw(frappe._("Save the document (with file attached) before Preview."), title=frappe._("Preview"))
		if not self.import_file:
			frappe.throw(frappe._("Attach a file first."), title=frappe._("Preview"))

		logger = frappe.logger("cheque_opening_import")
		logger.info("Preview start: doc=%s file_url=%s", self.name, self.import_file)

		self.set("items", [])
		valid = 0
		failed = 0

		try:
			rows = _get_sheet_from_file_url(self.import_file)
			parsed = _rows_from_sheet(rows)
			logger.info("Preview parsed: data_rows=%s", len(parsed))
		except Exception as e:
			logger.error("Preview parse failed: %s", e, exc_info=True)
			msg = cstr(e)
			self.append(
				"items",
				{
					"row_number": 0,
					"row_status": "Failed",
					"validation_message": "Parse failed",
					"error_detail": msg[:200],
				},
			)
			self.import_status = "Draft"
			self.summary = frappe._("Preview failed: {0}").format(msg)
			self.save()
			return {"total": 0, "valid": 0, "failed": 1, "summary": self.summary}

		if not parsed:
			self.import_status = "Draft"
			self.summary = frappe._("No data rows found in file.")
			self.append(
				"items",
				{
					"row_number": 0,
					"row_status": "Failed",
					"validation_message": "No data rows found in file.",
					"error_detail": "",
				},
			)
			self.save()
			return {"total": 0, "valid": 0, "failed": 1, "summary": self.summary}

		for row_no, row in parsed:
			err = _validate_row_preview(row, row_no)
			if err:
				failed += 1
				self.append(
					"items",
					{
						"row_number": row_no,
						"row_status": "Failed",
						"validation_message": err[:140],
						"error_detail": err[:200],
					},
				)
			else:
				valid += 1
				self.append(
					"items",
					{
						"row_number": row_no,
						"row_status": "Valid",
						"validation_message": f"OK → {cstr(row.get('workflow_state'))}",
					},
				)

		self.import_status = "Previewed"
		self.summary = frappe._("Preview: {0} data row(s), {1} valid, {2} failed.").format(
			len(parsed), valid, failed
		)
		self.save()
		return {"total": len(parsed), "valid": valid, "failed": failed, "summary": self.summary}

	@frappe.whitelist()
	def execute_import(self):
		self.check_permission("write")
		if not self.name:
			frappe.throw(frappe._("Save the document first."), title=frappe._("Import"))
		if not self.import_file:
			frappe.throw(frappe._("Attach a file first."), title=frappe._("Import"))
		if (self.import_status or "").strip() in ("Completed", "Completed With Errors"):
			frappe.throw(
				frappe._("This import has already been completed. Change the file and Preview again to re-run."),
				title=frappe._("Import"),
			)
		if (self.import_status or "").strip() != "Previewed":
			frappe.throw(
				frappe._("Run Preview first before Execute Import."),
				title=frappe._("Import"),
			)

		rows = _get_sheet_from_file_url(self.import_file)
		parsed = _rows_from_sheet(rows)

		# Guard: must have at least one valid row before executing.
		valid_candidates = 0
		for row_no, row in parsed:
			if not _validate_row_preview(row, row_no):
				valid_candidates += 1
				break
		if valid_candidates == 0:
			frappe.throw(
				frappe._("No valid rows found. Fix the file and run Preview again."),
				title=frappe._("Import"),
			)

		ok = 0
		bad = 0
		log_lines: list[str] = []

		# Pass current import doc id into row importer for traceability.
		frappe.flags.cheque_opening_import_name = self.name

		self.items = []
		for row_no, row in parsed:
			pre = _validate_row_preview(row, row_no)
			if pre:
				bad += 1
				log_lines.append(pre)
				self.append(
					"items",
					{
						"row_number": row_no,
						"row_status": "Failed",
						"validation_message": "Skipped",
						"error_detail": pre[:200],
					},
				)
				continue

			try:
				with savepoint():
					pdc_name = import_row(row_no, row)
				ok += 1
				self.append(
					"items",
					{
						"row_number": row_no,
						"row_status": "Imported",
						"validation_message": pdc_name,
						"imported_pdc": pdc_name,
						"post_dated_cheque": pdc_name,
					},
				)
				log_lines.append(f"Row {row_no}: OK → {pdc_name}")
			except Exception as e:
				bad += 1
				msg = f"Row {row_no}: {e!s}"
				log_lines.append(msg)
				frappe.log_error(title=f"Cheque Opening Import row {row_no}", message=frappe.get_traceback())
				self.append(
					"items",
					{
						"row_number": row_no,
						"row_status": "Failed",
						"validation_message": cstr(e)[:140],
						"error_detail": cstr(e)[:200],
					},
				)

		details = {
			"total": len(parsed),
			"imported": ok,
			"failed": bad,
			"lines": log_lines,
		}
		human = [
			f"Imported: {ok}",
			f"Failed: {bad}",
			f"Total: {len(parsed)}",
			"",
			"Details (JSON):",
			json.dumps(details, indent=2, default=str),
		]
		self.summary = "\n".join(human)
		self.import_status = "Completed" if bad == 0 else "Completed With Errors"
		self.save()

		return {"imported": ok, "failed": bad, "total": len(parsed), "summary": self.summary}


TEMPLATE_HEADERS: list[str] = [
	"cheque_direction",
	"company",
	"bank_account",
	"cheque_book",
	"cheque_number",
	"cheque_due_date",
	"cheque_amount",
	"party_type",
	"party",
	"workflow_state",
	"drawer_bank_name",
	"received_date",
	"sent_to_bank_date",
	"handover_date",
	"cleared_date",
	"bounced_date",
	"returned_date",
	"sayad_code",
]


@frappe.whitelist()
def download_import_template():
	"""Download a three-sheet **.xlsx** template (Data + Instructions + Allowed Values)."""
	if frappe.session.user == "Guest":
		frappe.throw(frappe._("Login required."), frappe.PermissionError)

	import openpyxl
	from openpyxl.styles import Font

	wb = openpyxl.Workbook()
	ws_data = wb.active
	ws_data.title = "Data"

	bold = Font(bold=True)
	ws_data.append(TEMPLATE_HEADERS)
	for cell in ws_data[1]:
		cell.font = bold

	# Sample A — Receivable, Draft (replace placeholders with real Link names from your site).
	ws_data.append(
		[
			CHEQUE_DIRECTION_RECEIVABLE,
			"__REPLACE_COMPANY__",
			"",
			"",
			"OB-REC-0001",
			date(2026, 6, 1),
			5000.0,
			"Customer",
			"__REPLACE_CUSTOMER__",
			"Draft",
			"__REPLACE_DRAWER_BANK__",
			date(2026, 1, 15),
			None,
			None,
			None,
			None,
			None,
			None,
		]
	)

	# Sample B — Payable, Registered, with book + number to match a Cheque Leaf.
	ws_data.append(
		[
			CHEQUE_DIRECTION_PAYABLE,
			"__REPLACE_COMPANY__",
			"__REPLACE_BANK_ACCOUNT__",
			"__REPLACE_CHEQUE_BOOK__",
			"101",
			date(2026, 6, 15),
			10000.0,
			"Supplier",
			"__REPLACE_SUPPLIER__",
			"Registered",
			"",
			date(2026, 1, 10),
			None,
			None,
			None,
			None,
			None,
			None,
		]
	)

	for col in range(1, len(TEMPLATE_HEADERS) + 1):
		ws_data.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

	ws_inst = wb.create_sheet("Instructions", 1)
	ws_inst.column_dimensions["A"].width = 110
	instructions = [
		"Cheque Opening Import — how to use this template",
		"",
		"Receivable: does NOT use cheque_leaf or cheque_book. Leave bank_account empty while in Draft; it becomes required for certain later workflow states (e.g. Sent to Bank). drawer_bank_name must be a valid Bank record (Link).",
		"",
		"Payable: you may set cheque_book + cheque_number to match an existing Cheque Leaf (same company and bank_account as the leaf). If the leaf is not found, the row will fail validation. This template does NOT create Cheque Books or leaves.",
		"",
		"Opening balance / GL: this tool only creates or updates Post Dated Cheque documents through the normal PDC save/submit path. It does NOT pick Opening Balance or Temporary Opening accounts here. Chart-of-accounts opening entries (including any auditor-required opening JEs) must be handled separately in Accounting unless a future enhancement adds dedicated import JEs.",
		"",
		"Journal entries: do not assume this import is silent in GL — existing PDC workflow rules may still create lifecycle Journal Entries when states change (per PDC Settings). Coordinate with finance so opening trial balance stays correct.",
		"",
		"Replace all __REPLACE_*__ placeholders with real values from your site (Company, Bank Account, Customer/Supplier, Cheque Book, Drawer Bank, etc.) before Preview / Execute.",
	]
	for line in instructions:
		ws_inst.append([line])

	ws_allow = wb.create_sheet("Allowed Values", 2)
	ws_allow.append(["Field", "Values / notes"])
	for c in ws_allow[1]:
		c.font = bold
	ws_allow.column_dimensions["A"].width = 28
	ws_allow.column_dimensions["B"].width = 90
	ws_allow.append(["cheque_direction", f"{CHEQUE_DIRECTION_RECEIVABLE}, {CHEQUE_DIRECTION_PAYABLE}"])
	ws_allow.append(
		[
			"workflow_state (examples; full set must be valid for the chosen direction)",
			", ".join(ALL_WORKFLOW_STATES),
		]
	)
	ws_allow.append(
		[
			"party_type (common)",
			"Customer, Supplier, Employee, Shareholder (must match Party master)",
		]
	)

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	frappe.local.response.filename = "cheque_opening_import_template.xlsx"
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "download"
